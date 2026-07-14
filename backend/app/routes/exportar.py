from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import tempfile
import os

from app.database import get_db
from app.models.models import Indicador, Riesgo7M, CriterioValoracion, CincoPorQue, CincoPorQuePregunta, AMFECRow, CostoBeneficio

router = APIRouter()

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
title_font = Font(name='Calibri', size=14, bold=True)
header_font = Font(name='Calibri', size=10, bold=True)
normal_font = Font(name='Calibri', size=10)
wrap_align = Alignment(wrap_text=True, vertical='top', horizontal='left')
center_align = Alignment(wrap_text=True, vertical='center', horizontal='center')


def set_cell(ws, row, col, value, font=normal_font, align=wrap_align, border=thin_border, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.alignment = align
    if border:
        cell.border = border
    if fill:
        cell.fill = fill


@router.get("/exportar/{indicador_id}")
def exportar_excel(indicador_id: int, db: Session = Depends(get_db)):
    ind = db.query(Indicador).get(indicador_id)
    if not ind:
        raise HTTPException(404, "Indicador no encontrado")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: Matriz 7M
    ws1 = wb.create_sheet("Matriz 7M")
    ws1.merge_cells('A1:F1')
    set_cell(ws1, 1, 1, f"INDICADOR: {ind.nombre}", title_font, center_align, None)
    ws1.merge_cells('A3:F3')
    set_cell(ws1, 3, 1, "MATRIZ 7M - Identificacion de Riesgos", header_font, center_align, thin_border, header_fill)
    headers = ['7M', 'Fuente', 'Riesgo', 'Descripcion', 'Efecto', 'Controles']
    for i, h in enumerate(headers):
        set_cell(ws1, 4, i + 1, h, header_font, center_align, thin_border, header_fill)
    riesgos = db.query(Riesgo7M).filter(Riesgo7M.indicador_id == indicador_id).all()
    for i, r in enumerate(riesgos):
        row = 5 + i
        for j, val in enumerate([r.categoria, r.fuente, r.riesgo, r.descripcion, r.efecto, r.controles]):
            set_cell(ws1, row, j + 1, val, normal_font, wrap_align, thin_border)
    ws1.column_dimensions['A'].width = 18
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 30
    ws1.column_dimensions['D'].width = 40
    ws1.column_dimensions['E'].width = 35
    ws1.column_dimensions['F'].width = 30

    # Sheet 2: Criterios
    ws2 = wb.create_sheet("Criterios de Valoracion")
    ws2.merge_cells('A1:D1')
    set_cell(ws2, 1, 1, "CRITERIOS DE VALORACION DE RIESGOS", title_font, center_align, None)
    ws2.merge_cells('A2:D2')
    set_cell(ws2, 2, 1, ind.nombre, header_font, center_align, None)
    for tipo_idx, tipo in enumerate(["PROBABILIDAD", "IMPACTO", "DETECTABILIDAD", "NPR"]):
        sr = 4 + tipo_idx * 8
        ws2.merge_cells(f'A{sr}:D{sr}')
        set_cell(ws2, sr, 1, f"{tipo_idx + 1}. CRITERIOS DE {tipo}", header_font, center_align, thin_border, header_fill)
        cols = {
            "PROBABILIDAD": ['Nivel', 'Rango', 'Descripcion', 'Referencia'],
            "IMPACTO": ['Nivel', 'Descripcion', 'Efecto en Indicador', 'Consecuencia Institucional'],
            "DETECTABILIDAD": ['Nivel', 'Descripcion', 'Capacidad de Deteccion', 'Ejemplo'],
            "NPR": ['Rango NPR', 'Nivel de Riesgo', 'Clasificacion', 'Accion Requerida'],
        }
        for i, h in enumerate(cols[tipo]):
            set_cell(ws2, sr + 1, i + 1, h, header_font, center_align, thin_border, PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'))
        criterios = db.query(CriterioValoracion).filter(CriterioValoracion.tipo == tipo).all()
        for i, c in enumerate(criterios):
            row = sr + 2 + i
            set_cell(ws2, row, 1, c.nivel, normal_font, center_align, thin_border)
            set_cell(ws2, row, 2, c.descripcion if tipo != "PROBABILIDAD" else c.referencia, normal_font, wrap_align, thin_border)
            set_cell(ws2, row, 3, c.referencia if tipo == "PROBABILIDAD" else (c.descripcion if tipo in ["IMPACTO", "DETECTABILIDAD"] else ""), normal_font, wrap_align, thin_border)
            set_cell(ws2, row, 4, c.descripcion if tipo == "DETECTABILIDAD" else "", normal_font, wrap_align, thin_border)
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 35
    ws2.column_dimensions['D'].width = 40

    # Sheet 3: 5 Por Que
    ws3 = wb.create_sheet("5 Por Que")
    ws3.merge_cells('A1:B1')
    set_cell(ws3, 1, 1, "ANALISIS DE LOS 5 POR QUE", title_font, center_align, None)
    ws3.merge_cells('A2:B2')
    set_cell(ws3, 2, 1, ind.nombre, header_font, center_align, None)
    r = 4
    analisis_list = db.query(CincoPorQue).filter(CincoPorQue.indicador_id == indicador_id).all()
    for analisis in analisis_list:
        ws3.merge_cells(f'A{r}:B{r}')
        set_cell(ws3, r, 1, analisis.titulo, header_font, wrap_align, thin_border, header_fill)
        r += 1
        set_cell(ws3, r, 1, "Pregunta", header_font, center_align, thin_border, PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'))
        set_cell(ws3, r, 2, "Respuesta", header_font, center_align, thin_border, PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'))
        r += 1
        preguntas = db.query(CincoPorQuePregunta).filter(CincoPorQuePregunta.analisis_id == analisis.id).order_by(CincoPorQuePregunta.nivel).all()
        for p in preguntas:
            set_cell(ws3, r, 1, p.pregunta, normal_font, wrap_align, thin_border)
            set_cell(ws3, r, 2, p.respuesta, normal_font, wrap_align, thin_border)
            r += 1
        ws3.merge_cells(f'A{r}:B{r}')
        set_cell(ws3, r, 1, f"Causa Raiz: {analisis.causa_raiz}", Font(name='Calibri', size=10, bold=True, italic=True), wrap_align, thin_border)
        r += 2
    ws3.column_dimensions['A'].width = 45
    ws3.column_dimensions['B'].width = 60

    # Sheet 4: AMFEC
    ws4 = wb.create_sheet("AMFEC")
    ws4.merge_cells('A1:P1')
    set_cell(ws4, 1, 1, f"AMFEC - {ind.nombre}", title_font, center_align, None)
    amfec_headers = ['7M', 'Componente', 'Funcion', 'Modo de Fallo', 'Efecto', 'Causa', 'Controles',
                     'Prob', 'Imp', 'Det', 'NPR', 'Clasif', 'Tratamiento', 'Accion', 'Responsable', 'Actividades']
    for i, h in enumerate(amfec_headers):
        set_cell(ws4, 3, i + 1, h, header_font, center_align, thin_border, header_fill)
    amfec_rows = db.query(AMFECRow).filter(AMFECRow.indicador_id == indicador_id).all()
    for i, row in enumerate(amfec_rows):
        r = 4 + i
        vals = [row.siete_m, row.componente, row.funcion, row.modo_fallo, row.efecto, row.causa, row.controles,
                row.probabilidad, row.impacto, row.detectabilidad, row.npr, row.clasificacion, row.tratamiento, row.accion, row.responsable, row.actividades]
        for j, val in enumerate(vals):
            set_cell(ws4, r, j + 1, val, normal_font, center_align if j in [7, 8, 9, 10, 11] else wrap_align, thin_border)
    widths = [14, 16, 25, 28, 28, 30, 25, 8, 8, 8, 8, 14, 14, 35, 20, 30]
    for i, w in enumerate(widths):
        ws4.column_dimensions[chr(65 + i)].width = w

    # Sheet 5: Costo Beneficio
    ws5 = wb.create_sheet("Costo Beneficio")
    ws5.merge_cells('A1:Q1')
    set_cell(ws5, 1, 1, "ANALISIS COSTO BENEFICIO", title_font, center_align, None)
    ws5.merge_cells('A2:Q2')
    set_cell(ws5, 2, 1, ind.nombre, header_font, center_align, None)
    cb_headers = ['#', 'Accion', 'Riesgo', 'Presup', 'Dism Imp', 'Dism Prob', 'Tiempo', 'Personal',
                  'Total B', 'Imagen', 'Result', 'Equipos', 'Legales', 'Prob Antes', 'Total C', 'B/C', 'Decision']
    for i, h in enumerate(cb_headers):
        set_cell(ws5, 4, i + 1, h, header_font, center_align, thin_border, header_fill)
    cb_rows = db.query(CostoBeneficio).filter(CostoBeneficio.indicador_id == indicador_id).all()
    for i, cb in enumerate(cb_rows):
        r = 5 + i
        vals = [i + 1, cb.accion, cb.riesgo_asociado, cb.presupuesto, cb.disminucion_impacto,
                cb.disminucion_probabilidad, cb.tiempo_implementacion, cb.personal_propio,
                cb.total_beneficios, cb.impacto_imagen, cb.impacto_resultados, cb.impacto_equipos,
                cb.impacto_legales, cb.probabilidad_antes, cb.total_consecuencias,
                cb.relacion_bc if cb.relacion_bc is not None else "", cb.decision]
        for j, val in enumerate(vals):
            set_cell(ws5, r, j + 1, val, normal_font, center_align if j in [0, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15] else wrap_align, thin_border)
    for i, w in enumerate([5, 30, 25, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 30]):
        ws5.column_dimensions[chr(65 + i)].width = w

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    return FileResponse(
        tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"riesgos_{ind.nombre[:30]}.xlsx"
    )
